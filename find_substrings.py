import openpyxl as o
import pandas as pd

hh_filename = 'hh_rabbitmq_moscow_positions.xlsx'
vk_filename = "habr_career__vk_jobs.xlsx"

wb = o.load_workbook(vk_filename)
work_sheet = wb['Sheet1']

result_entries = []

# for row in work_sheet.iter_rows(min_row=2, max_row=work_sheet.max_row):
#     for cell in row:
#         result_entries.append(cell.value)

for cell in work_sheet['A']:
    result_entries.append(cell.value.lower())  # to lower case

result_entries = result_entries[1:]  # remove 'name' field

list_of_key_words = ['php', 'scientist', 'data engineer',
                     'инженер данных', 'тестир', 'auto',
                     'авто', 'devops', 'go', 'mysql', 'python',
                     'c#', 'java', 'node', 'full', 'js', 'ETL', 'etl',
                     ]

positions = ['senior', 'lead', 'junior', 'middle', 'техлид', 'начинающий', 'главный', 'старший']


def count_positions(entries: list, input_list: list) -> pd.DataFrame:
    d = dict.fromkeys(entries, 0)

    for clause in input_list:
        for key_word in entries:
            if key_word in clause:
                d[key_word] += 1

    sorted_d = dict(sorted(d.items(), key=lambda var: var[1], reverse=True))

    output = pd.DataFrame(
        {
            "Key words": sorted_d.keys(),
            "Count": sorted_d.values()
        }
    )

    return output


def dump_to_excel(some_results, title: str):
    some_results.to_excel(f'{title}.xlsx', index=False)


hh_title = "hh_rabbitmq_moscow_positions_cleaned_data"


if __name__ == '__main__':
    result_key_words = count_positions(list_of_key_words, result_entries)
    dump_to_excel(result_key_words, "habr__vk_jobs")
